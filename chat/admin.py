from django.contrib import admin
from django.contrib.auth.admin import UserAdmin
from django.contrib.auth.models import User
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .models import (
    Appointment,
    Conversation,
    DataExportJob,
    FicheAttachment,
    FicheConsultation,
    FicheMessage,
    FicheReference,
    LabResult,
    MedecinAvailability,
    MedecinException,
    MessageIA,
    WebhookEvent,
)


@admin.register(Conversation)
class ConversationAdmin(admin.ModelAdmin):
    list_display = ("id", "user", "fiche", "created_at", "titre", "nom")
    list_filter = ("created_at", "fiche")
    search_fields = ("user__username", "fiche__numero_dossier")
    date_hierarchy = "created_at"
    ordering = ("-created_at",)


@admin.register(MessageIA)
class MessageIAAdmin(admin.ModelAdmin):
    list_display = ("id", "conversation", "role", "short_content", "timestamp")
    list_filter = ("role", "timestamp")
    search_fields = ("conversation__user__username", "content")
    date_hierarchy = "timestamp"
    ordering = ("-timestamp",)

    def short_content(self, obj):
        return (obj.content[:75] + "...") if len(obj.content) > 75 else obj.content

    short_content.short_description = "Contenu"


@admin.register(FicheConsultation)
class FicheConsultationAdmin(admin.ModelAdmin):
    list_display = ("id", "nom", "prenom", "date_naissance", "age", "numero_dossier", "status", "date_consultation", "conversations_count")
    list_filter = ("date_consultation", "status", "sexe", "etat_civil", "is_patient_distance")
    search_fields = ("nom", "prenom", "numero_dossier", "telephone")
    date_hierarchy = "date_consultation"
    ordering = ("-date_consultation",)
    actions = ["export_to_excel"]

    def conversations_count(self, obj):
        return obj.conversations.count()

    conversations_count.short_description = "Nb conversations"

    @admin.action(description="📊 Exporter les fiches sélectionnées en Excel")
    def export_to_excel(self, request, queryset):
        """Exporte les fiches de consultation sélectionnées vers un fichier Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Fiches Consultation"

        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Colonnes à exporter
        columns = [
            ("N° Dossier", "numero_dossier"),
            ("Nom", "nom"),
            ("Postnom", "postnom"),
            ("Prénom", "prenom"),
            ("Date Naissance", "date_naissance"),
            ("Âge", "age"),
            ("Sexe", "sexe"),
            ("Téléphone", "telephone"),
            ("État Civil", "etat_civil"),
            ("Occupation", "occupation"),
            ("Avenue", "avenue"),
            ("Quartier", "quartier"),
            ("Commune", "commune"),
            ("Contact Nom", "contact_nom"),
            ("Contact Tél", "contact_telephone"),
            ("Date Consultation", "date_consultation"),
            ("Statut", "status"),
            ("Température", "temperature"),
            ("SpO2", "spo2"),
            ("Poids", "poids"),
            ("Tension", "tension_arterielle"),
            ("Pouls", "pouls"),
            ("Motif Consultation", "motif_consultation"),
            ("Histoire Maladie", "histoire_maladie"),
            ("État Général", "etat"),
            ("Capacité Physique", "capacite_physique"),
            ("Capacité Psycho", "capacite_psychologique"),
            ("Fébrile", "febrile"),
            ("Diagnostic IA", "diagnostic_ia"),
            ("Diagnostic Médecin", "diagnostic"),
            ("Traitement", "traitement"),
            ("Recommandations", "recommandations"),
            ("Médecin Validateur", "medecin_validateur"),
            ("Date Validation", "date_validation"),
            ("Consultation Distance", "is_patient_distance"),
        ]

        # En-têtes
        for col_num, (header, _) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Données
        for row_num, fiche in enumerate(queryset, 2):
            for col_num, (_, field) in enumerate(columns, 1):
                value = getattr(fiche, field, "")
                
                # Gestion des ForeignKey
                if field == "medecin_validateur" and value:
                    value = f"{value.first_name} {value.last_name}".strip() or value.username
                
                # Gestion des dates
                if hasattr(value, "strftime"):
                    value = value.strftime("%d/%m/%Y %H:%M") if hasattr(value, "hour") else value.strftime("%d/%m/%Y")
                
                # Gestion des booléens
                if isinstance(value, bool):
                    value = "Oui" if value else "Non"
                
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        # Ajustement largeur colonnes
        for col_num, (header, _) in enumerate(columns, 1):
            column_letter = get_column_letter(col_num)
            max_length = len(header)
            for row in range(2, min(queryset.count() + 2, 50)):  # Limiter à 50 lignes pour le calcul
                cell_value = ws.cell(row=row, column=col_num).value
                if cell_value:
                    max_length = max(max_length, min(len(str(cell_value)), 50))
            ws.column_dimensions[column_letter].width = max_length + 2

        # Figer la première ligne
        ws.freeze_panes = "A2"

        # Réponse HTTP
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"fiches_consultation_{queryset.count()}_export.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)

        self.message_user(request, f"✅ {queryset.count()} fiche(s) exportée(s) avec succès!")
        return response


@admin.register(Appointment)
class AppointmentAdmin(admin.ModelAdmin):
    list_display = ("id", "patient", "medecin", "status", "requested_start", "confirmed_start", "created_at")
    list_filter = ("status", "created_at")
    search_fields = ("patient__username", "medecin__username")
    date_hierarchy = "created_at"
    ordering = ("-created_at",)


@admin.register(FicheMessage)
class FicheMessageAdmin(admin.ModelAdmin):
    list_display = ("id", "fiche", "author", "short_content", "created_at")
    list_filter = ("created_at",)
    search_fields = ("fiche__numero_dossier", "author__username", "content")
    date_hierarchy = "created_at"
    ordering = ("-created_at",)

    def short_content(self, obj):
        return (obj.content[:60] + "...") if len(obj.content) > 60 else obj.content


@admin.register(FicheReference)
class FicheReferenceAdmin(admin.ModelAdmin):
    list_display = ("id", "fiche", "title", "source", "year", "created_at")
    list_filter = ("source", "year", "created_at")
    search_fields = ("title", "authors", "fiche__numero_dossier")
    date_hierarchy = "created_at"
    ordering = ("-created_at",)


@admin.register(LabResult)
class LabResultAdmin(admin.ModelAdmin):
    list_display = ("id", "fiche", "type_analyse", "valeur", "unite", "date_prelevement", "laboratoire")
    list_filter = ("type_analyse", "date_prelevement", "laboratoire")
    search_fields = ("fiche__numero_dossier", "type_analyse", "laboratoire")
    date_hierarchy = "date_prelevement"
    ordering = ("-date_prelevement",)


@admin.register(FicheAttachment)
class FicheAttachmentAdmin(admin.ModelAdmin):
    list_display = ("id", "fiche", "kind", "file", "uploaded_by", "created_at")
    list_filter = ("kind", "created_at")
    search_fields = ("fiche__numero_dossier", "note", "uploaded_by__username")
    date_hierarchy = "created_at"
    ordering = ("-created_at",)


@admin.register(MedecinAvailability)
class MedecinAvailabilityAdmin(admin.ModelAdmin):
    list_display = ("id", "medecin", "get_day_display", "start_time", "end_time", "consultation_type", "is_active")
    list_filter = ("day_of_week", "consultation_type", "is_active", "created_at")
    search_fields = ("medecin__username", "medecin__first_name", "medecin__last_name")
    ordering = ("medecin", "day_of_week", "start_time")

    def get_day_display(self, obj):
        return obj.get_day_of_week_display()

    get_day_display.short_description = "Jour"


@admin.register(MedecinException)
class MedecinExceptionAdmin(admin.ModelAdmin):
    list_display = ("id", "medecin", "exception_type", "start_datetime", "end_datetime", "is_recurring")
    list_filter = ("exception_type", "is_recurring", "created_at")
    search_fields = ("medecin__username", "reason")
    date_hierarchy = "start_datetime"
    ordering = ("-start_datetime",)


@admin.register(WebhookEvent)
class WebhookEventAdmin(admin.ModelAdmin):
    list_display = ("id", "event_type", "sender_phone", "processing_status", "related_user", "received_at")
    list_filter = ("event_type", "processing_status", "received_at")
    search_fields = ("sender_phone", "recipient_phone", "external_id", "content")
    date_hierarchy = "received_at"
    ordering = ("-received_at",)
    readonly_fields = ("received_at", "processed_at")

    fieldsets = (
        ("Informations Webhook", {"fields": ("event_type", "external_id", "sender_phone", "recipient_phone")}),
        ("Contenu", {"fields": ("content", "raw_payload")}),
        ("Traitement", {"fields": ("processing_status", "processing_error", "received_at", "processed_at")}),
        ("Associations", {"fields": ("related_user", "related_fiche", "created_message")}),
    )


@admin.register(DataExportJob)
class DataExportJobAdmin(admin.ModelAdmin):
    list_display = ("id", "created_by", "export_format", "date_range", "status", "records_count", "created_at")
    list_filter = ("export_format", "status", "include_personal_data", "created_at")
    search_fields = ("created_by__username",)
    date_hierarchy = "created_at"
    ordering = ("-created_at",)
    readonly_fields = (
        "status",
        "file_path",
        "file_size",
        "records_count",
        "error_message",
        "started_at",
        "completed_at",
        "created_at",
    )

    def date_range(self, obj):
        return f"{obj.date_start} → {obj.date_end}"

    date_range.short_description = "Période"

    fieldsets = (
        (
            "Paramètres Export",
            {"fields": ("created_by", "export_format", "date_start", "date_end", "include_personal_data", "filters")},
        ),
        ("Résultats", {"fields": ("status", "file_path", "file_size", "records_count", "error_message")}),
        ("Timing", {"fields": ("created_at", "started_at", "completed_at")}),
    )
